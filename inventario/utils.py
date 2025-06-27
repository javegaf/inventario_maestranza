""" Utilidades para la generación de informes de inventario en PDF y Excel. """
from datetime import datetime
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import Producto, OrdenCompra, ItemOrdenCompra
from django.db import transaction
from django.db.models import F

def exportar_pdf_inventario(productos, request):
    """
    Genera un archivo PDF del informe de inventario filtrado.
    """
    template_path = 'reportes/informe_pdf.html'  # Asegúrate de tener esta plantilla
    context = {
        'productos': productos,
        'usuario': request.user,
        'fecha': datetime.now(),
        'filtros': request.GET.dict()
    }
    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="informe_inventario.pdf"'
    pisa.CreatePDF(BytesIO(html.encode("UTF-8")), dest=response, encoding='UTF-8')
    return response

def exportar_excel_inventario(productos):
    """
    Genera un archivo Excel del informe de inventario filtrado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Cabeceras
    headers = ['Nombre', 'Ubicación', 'Categoría', 'Stock Actual', 'Stock Mínimo']
    ws.append(headers)

    # Estilo de cabeceras
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Filas de productos
    for p in productos:
        ws.append([
            p.nombre,
            p.ubicacion,
            p.categoria,
            p.stock_actual,
            p.stock_minimo,
        ])

    # Autoajustar ancho de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Exportar a HTTP response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="informe_inventario.xlsx"'
    return response



@transaction.atomic
def generar_ordenes_sugeridas():
    productos_bajo_stock = Producto.objects.filter(stock_actual__lt=F('stock_minimo'))

    for producto in productos_bajo_stock:
        proveedor = producto.proveedor
        cantidad_faltante = producto.stock_minimo - producto.stock_actual

        # Buscar orden sugerida abierta para el proveedor
        orden = OrdenCompra.objects.filter(
            proveedor=proveedor,
            estado='sugerida'
        ).first()

        if not orden:
            orden = OrdenCompra.objects.create(
                proveedor=proveedor,
                estado='sugerida',
                observaciones='Orden generada automáticamente por stock bajo.'
            )

        # Verificar si el producto ya está en la orden
        item_existente = ItemOrdenCompra.objects.filter(
            orden=orden,
            producto=producto
        ).first()

        if not item_existente:
            ItemOrdenCompra.objects.create(
                orden=orden,
                producto=producto,
                cantidad=cantidad_faltante
            )
